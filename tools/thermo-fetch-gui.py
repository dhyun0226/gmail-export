"""
Thermo Fisher 카탈로그 조회 (GUI 버전).

비개발자용 더블클릭 실행 EXE 빌드 대상.
빌드:  pyinstaller --onefile --windowed --name thermo-fetch thermo-fetch-gui.py
"""
from __future__ import annotations

import datetime as dt
import html
import random
import re
import threading
import time
import tkinter as tk
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import requests
from openpyxl import Workbook, load_workbook

PDP_URL = "https://www.thermofisher.com/order/catalog/product/{cat}"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
}
RX_NAME = re.compile(r'<h1[^>]*role="heading"[^>]*>([^<]+)</h1>')
RX_DESC = re.compile(r'<div class="short-description">([^<]+)</div>')
MISSING = "없음"


def clean(s: str | None) -> str:
    return html.unescape(s).strip() if s else ""


def _try(session: requests.Session, cat: str) -> tuple[str, str, str] | None:
    """단일 시도. 성공시 (name, desc, url), 실패시 None."""
    url = PDP_URL.format(cat=cat)
    try:
        r = session.get(url, headers=HEADERS, timeout=25, allow_redirects=True)
    except requests.RequestException:
        return None
    if r.status_code != 200:
        return None
    body = r.text
    m_name = RX_NAME.search(body)
    if not m_name:
        return None
    name = clean(m_name.group(1))
    if not name:
        return None
    m_desc = RX_DESC.search(body)
    desc = clean(m_desc.group(1) if m_desc else "")
    return name, desc, url


def _candidates(catalog: str) -> list[str]:
    """엑셀 자동 숫자변환으로 앞 0이 잘린 경우 대비, 6자리 zfill 폴백 추가."""
    cands = [catalog]
    if catalog.isdigit() and 1 <= len(catalog) < 6:
        cands.append(catalog.zfill(6))
    return cands


def fetch_one(session: requests.Session, catalog: str) -> tuple[str, str, str, str]:
    """Returns (catalog, name, description, url)."""
    for cat in _candidates(catalog):
        hit = _try(session, cat)
        if hit:
            name, desc, url = hit
            return cat, name, desc, url
    return catalog, MISSING, MISSING, PDP_URL.format(cat=catalog)


def load_catalogs(path: Path) -> list[str]:
    """엑셀 A열에서 카탈로그 번호 로드. 헤더('카탈로그 번호' 등)는 자동 스킵."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    out: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row:
            continue
        v = row[0]
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        # 헤더 자동 스킵 (한국어/영어)
        if i == 0 and not re.match(r"^[A-Za-z0-9_.\-]+$", s):
            continue
        out.append(s)
    return out


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Thermo Fisher 카탈로그 조회")
        self.geometry("680x520")
        self.resizable(False, False)

        self.input_path: Path | None = None
        self.cancel_flag = threading.Event()
        self.worker: threading.Thread | None = None

        self._build_ui()

    def _build_ui(self) -> None:
        pad = {"padx": 12, "pady": 6}

        ttk.Label(
            self,
            text="Thermo Fisher 카탈로그 조회",
            font=("Malgun Gothic", 14, "bold"),
        ).pack(anchor="w", **pad)

        guide = (
            "1) 양식 엑셀의 A열에 카탈로그 번호를 입력 (1행은 헤더)\n"
            "2) [파일 선택] → 양식 파일 선택\n"
            "3) [시작] 클릭 → 같은 폴더에 결과 엑셀 생성\n"
            "※ 등록되지 않은 카탈로그는 이름/설명에 \"없음\"으로 표시됩니다."
        )
        ttk.Label(self, text=guide, justify="left").pack(anchor="w", **pad)

        frm = ttk.Frame(self)
        frm.pack(fill="x", **pad)
        self.path_var = tk.StringVar(value="(선택 안 됨)")
        ttk.Button(frm, text="파일 선택", command=self.choose).pack(side="left")
        ttk.Label(frm, textvariable=self.path_var, foreground="#555").pack(
            side="left", padx=10
        )

        opt = ttk.Frame(self)
        opt.pack(fill="x", **pad)
        ttk.Label(opt, text="동시 요청 수:").pack(side="left")
        self.workers_var = tk.IntVar(value=4)
        ttk.Spinbox(
            opt, from_=1, to=10, textvariable=self.workers_var, width=5
        ).pack(side="left", padx=6)

        self.start_btn = ttk.Button(self, text="시작", command=self.start)
        self.start_btn.pack(anchor="w", **pad)
        self.stop_btn = ttk.Button(
            self, text="중지", command=self.stop, state="disabled"
        )
        self.stop_btn.pack(anchor="w", padx=12)

        self.progress = ttk.Progressbar(self, length=640, mode="determinate")
        self.progress.pack(**pad)
        self.status_var = tk.StringVar(value="대기 중")
        ttk.Label(self, textvariable=self.status_var).pack(anchor="w", **pad)

        self.log = tk.Text(self, height=14, font=("Consolas", 9))
        self.log.pack(fill="both", expand=True, padx=12, pady=6)

    def log_line(self, msg: str) -> None:
        self.log.insert("end", msg + "\n")
        self.log.see("end")

    def choose(self) -> None:
        p = filedialog.askopenfilename(
            title="입력 양식 엑셀 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if p:
            self.input_path = Path(p)
            self.path_var.set(str(self.input_path))

    def start(self) -> None:
        if not self.input_path:
            messagebox.showwarning("안내", "먼저 입력 엑셀 파일을 선택하세요.")
            return
        try:
            catalogs = load_catalogs(self.input_path)
        except Exception as e:  # noqa: BLE001
            messagebox.showerror("파일 오류", f"엑셀을 읽을 수 없습니다: {e}")
            return
        if not catalogs:
            messagebox.showwarning("안내", "A열에서 카탈로그 번호를 찾지 못했습니다.")
            return

        self.cancel_flag.clear()
        self.start_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.progress.config(maximum=len(catalogs), value=0)
        self.log.delete("1.0", "end")
        self.log_line(f"총 {len(catalogs)}건 처리 시작")

        self.worker = threading.Thread(
            target=self._run, args=(catalogs,), daemon=True
        )
        self.worker.start()

    def stop(self) -> None:
        self.cancel_flag.set()
        self.log_line("중지 요청…")

    def _run(self, catalogs: list[str]) -> None:
        results: list[tuple[str, str, str, str] | None] = [None] * len(catalogs)
        ok = 0
        miss = 0
        workers = max(1, min(10, self.workers_var.get()))
        session = requests.Session()
        started = time.time()

        def task(idx: int, cat: str) -> tuple[int, tuple[str, str, str, str]]:
            if self.cancel_flag.is_set():
                return idx, (cat, MISSING, MISSING, PDP_URL.format(cat=cat))
            time.sleep(random.uniform(0.3, 0.8))
            return idx, fetch_one(session, cat)

        try:
            with ThreadPoolExecutor(max_workers=workers) as ex:
                futs = [ex.submit(task, i, c) for i, c in enumerate(catalogs)]
                done = 0
                for fut in as_completed(futs):
                    idx, row = fut.result()
                    results[idx] = row
                    done += 1
                    if row[1] == MISSING:
                        miss += 1
                        mark = "없음"
                    else:
                        ok += 1
                        mark = "OK"
                    self.after(
                        0,
                        self._tick,
                        done,
                        len(catalogs),
                        ok,
                        miss,
                        row[0],
                        row[1],
                        mark,
                        started,
                    )
                    if self.cancel_flag.is_set():
                        # 남은 future들도 task() 안에서 자동으로 MISSING 처리
                        pass
        except Exception as e:  # noqa: BLE001
            self.after(0, messagebox.showerror, "오류", str(e))
            self.after(0, self._finish_ui)
            return

        # 미처리 슬롯 채우기 (중지 시 대비)
        for i, r in enumerate(results):
            if r is None:
                results[i] = (catalogs[i], MISSING, MISSING, PDP_URL.format(cat=catalogs[i]))

        out_path = self.input_path.parent / (
            "결과_" + dt.datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "결과"
        ws.append(["카탈로그 번호", "이름", "설명", "URL"])
        for r in results:
            ws.append(list(r))
        for col, w in zip("ABCD", [16, 60, 80, 60]):
            ws.column_dimensions[col].width = w
        wb.save(out_path)

        elapsed = time.time() - started
        self.after(
            0,
            self.log_line,
            f"완료. ok={ok} 없음={miss} 시간={elapsed:.1f}s -> {out_path}",
        )
        self.after(0, self.status_var.set, f"완료: {out_path}")
        self.after(0, self._finish_ui)
        self.after(0, lambda: messagebox.showinfo("완료", f"결과 저장:\n{out_path}"))

    def _tick(
        self,
        done: int,
        total: int,
        ok: int,
        miss: int,
        cat: str,
        name: str,
        mark: str,
        started: float,
    ) -> None:
        self.progress.config(value=done)
        elapsed = time.time() - started
        rate = done / elapsed if elapsed > 0 else 0
        eta = (total - done) / rate if rate > 0 else 0
        self.status_var.set(
            f"{done}/{total}  성공:{ok}  없음:{miss}  속도:{rate:.1f}/s  남은시간:{int(eta)}s"
        )
        self.log_line(f"[{done}/{total}] {mark} {cat}  {name[:60]}")

    def _finish_ui(self) -> None:
        self.start_btn.config(state="normal")
        self.stop_btn.config(state="disabled")


if __name__ == "__main__":
    App().mainloop()
